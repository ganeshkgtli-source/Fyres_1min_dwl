import pandas as pd
import time
import io
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from accounts.services.fyers_client import get_fyers_client
from accounts.services.symbol_service import download_bse_symbols, get_filtered_symbols
from accounts.services.data_formatter import convert_time

from accounts.models import BhavcopyFile, OneMinDataFile


# =========================================================
# FORMAT DATAFRAME
# =========================================================

def format_dataframe(df, symbol):

    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp"])

    df["SYMBOL"] = symbol
    df["TIME_FRAME"] = "1_MIN"

    df["DATE"] = df["timestamp"].dt.date
    df["TIME"] = df["timestamp"].dt.time
    df["DAY"] = df["timestamp"].dt.day_name().str.upper()

    market_open = 9 * 60 + 15

    def candle(ts):
        minutes = ts.hour * 60 + ts.minute
        diff = minutes - market_open
        return 0 if diff < 0 else diff + 1

    df["CANDLE_COUNT"] = df["timestamp"].apply(candle)

    df["CATEGORY"] = "XX"

    df = df.rename(columns={
        "open": "OPEN",
        "high": "HIGH",
        "low": "LOW",
        "close": "CLOSE",
        "volume": "VOLUME"
    })

    df["TOTAL_VOLUME"] = df.groupby("DATE")["VOLUME"].cumsum()

    df["SHARE_CHECK"] = ""

    return df[[
        "DATE","DAY","TIME_FRAME","CATEGORY","SYMBOL",
        "CANDLE_COUNT","TIME","OPEN","HIGH","LOW",
        "CLOSE","VOLUME","TOTAL_VOLUME","SHARE_CHECK"
    ]]


# =========================================================
# DATE RANGE CHUNKS
# =========================================================

def daterange_chunks(start, end, chunk=100):

    curr = start

    while curr < end:
        next_date = min(curr + timedelta(days=chunk-1), end)
        yield curr, next_date
        curr = next_date + timedelta(days=1)


# =========================================================
# APPEND EOD ROWS (FROM DB)
# =========================================================

def append_eod_rows(df, symbol):

    if df.empty:
        return df

    bse_symbol = symbol.strip()
    eod_time = datetime.strptime("15:30:59","%H:%M:%S").time()

    eod_rows = []

    for date in sorted(df["DATE"].unique()):

        date_obj = pd.to_datetime(date).date()

        try:
            bhav_obj = BhavcopyFile.objects.filter(
                trade_date=date_obj,
                is_deleted=False
            ).first()

            if not bhav_obj:
                continue

            bhav = pd.read_csv(io.BytesIO(bhav_obj.file_data))
            bhav.columns = bhav.columns.str.strip().str.upper()

        except Exception as e:
            print("BHAV DB ERROR:", e)
            continue

        bhav["SYMBOL"] = bhav["SYMBOL"].astype(str).str.strip()

        row = bhav[bhav["SYMBOL"] == bse_symbol]

        if row.empty:
            continue

        r = row.iloc[0]

        try:

            day_df = df[df["DATE"] == date].sort_values("TIME")

            last_close = None
            last_total = None

            if not day_df.empty:
                last_close = float(day_df.iloc[-1]["CLOSE"])
                last_total = float(day_df.iloc[-1]["TOTAL_VOLUME"])

            eod_close  = float(r["CLOSE"])
            eod_volume = float(r["NO_OF_SHRS"])

            close_diff = None
            volume_diff = None

            if last_close is not None:
                close_diff = round(eod_close - last_close,4)

            if last_total is not None:
                volume_diff = int(eod_volume - last_total)

            # FIX CLOSE
            if last_close is not None and eod_close != last_close:
                eod_close = last_close

            # FIX VOLUME
            if last_total is not None and eod_volume != last_total:
                eod_volume = last_total

            share_check = "YES" if last_total == eod_volume else "NO"

            eod_rows.append({
                "DATE":date,
                "DAY":pd.Timestamp(date_obj).day_name().upper(),
                "TIME_FRAME":"EOD",
                "CATEGORY":"XX",
                "SYMBOL":symbol,
                "CANDLE_COUNT":999,
                "TIME":eod_time,

                "OPEN":float(r["OPEN"]),
                "HIGH":float(r["HIGH"]),
                "LOW":float(r["LOW"]),

                "CLOSE":eod_close,
                "VOLUME":eod_volume,
                "TOTAL_VOLUME":eod_volume,

                "CLOSE_DIFF":close_diff,
                "VOLUME_DIFF":volume_diff,

                "SHARE_CHECK":share_check
            })

        except Exception as e:
            print("EOD ERROR:", e)

    if not eod_rows:
        return df

    eod_df = pd.DataFrame(eod_rows)

    final_df = pd.concat([df,eod_df],ignore_index=True)
    final_df = final_df.sort_values(["DATE","TIME"]).reset_index(drop=True)

    return final_df


# =========================================================
# MAIN DOWNLOAD FUNCTION
# =========================================================

def download_1min_data(client_id):

    print("\n==============================")
    print("STARTING 1 MIN DATA DOWNLOAD")
    print("==============================")

    fyers = get_fyers_client(client_id)

    print("\nFetching symbol list...")

    df_symbols = download_bse_symbols()
    symbols = get_filtered_symbols(df_symbols)

    print("Total symbols:",len(symbols))

    start_date = datetime(2018,4,1)
    now = datetime.now()

    # Before 4 PM → previous day
    if now.hour < 16:
        end_date = now - timedelta(days=1)
    else:
        end_date = now

    api_calls = 0

    for symbol in symbols:

        print("\n===================================")
        print("Downloading:",symbol)
        print("===================================")

        main_df = pd.DataFrame()
        rotation = 0

        for chunk_start,chunk_end in daterange_chunks(start_date,end_date):

            print("ROTATION",rotation,"|",chunk_start.date(),"->",chunk_end.date())

            payload = {
                "symbol":symbol,
                "resolution":"1",
                "date_format":"1",
                "range_from":chunk_start.strftime("%Y-%m-%d"),
                "range_to":chunk_end.strftime("%Y-%m-%d"),
                "cont_flag":"1"
            }

            try:
                history = fyers.history(payload)
                api_calls += 1

                if "candles" in history and history["candles"]:

                    df = pd.DataFrame(
                        history["candles"],
                        columns=["timestamp","open","high","low","close","volume"]
                    )

                    df = convert_time(df)
                    df = format_dataframe(df,symbol)

                    main_df = pd.concat([main_df,df],ignore_index=True)

                    print("OK:",len(df),"rows")

                else:
                    print("WARNING: No candles")

            except Exception as e:
                print("ERROR:",e)

            rotation += 1
            time.sleep(0.3)

        

        if main_df.empty:
            print("WARNING: No data for",symbol)

        else:
            main_df = main_df.drop_duplicates(
                subset=["DATE","TIME","SYMBOL"]
            ).reset_index(drop=True)

            print("Appending EOD rows")
            main_df = append_eod_rows(main_df,symbol)

        
        # SAVE TO DB
        

        if not main_df.empty:

            file_symbol = symbol.replace(":","_").replace("-","_")
            file_name = f"{file_symbol}_1MIN.xlsx"

            # Save to memory
            output = io.BytesIO()
            main_df.to_excel(output, index=False, engine="openpyxl")
            output.seek(0)

            # Highlight EOD
            wb = load_workbook(output)
            ws = wb.active

            highlight = PatternFill(
                start_color="90EE90",
                end_color="90EE90",
                fill_type="solid"
            )

            time_frame_col = None

            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "TIME_FRAME":
                    time_frame_col = col
                    break

            if time_frame_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    if str(row[time_frame_col - 1].value).strip() == "EOD":
                        for cell in row:
                            cell.fill = highlight

            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            wb.close()
            final_buffer.seek(0)

            # Store in DB
            OneMinDataFile.objects.update_or_create(
                symbol=symbol,
                file_name=file_name,
                defaults={
                    "file_data": final_buffer.getvalue(),
                    "is_deleted": False
                }
            )

            print("Saved to DB:",file_name)

        else:
            print("SKIPPED:",symbol)

        time.sleep(1)

    print("\n====================================")
    print("DOWNLOAD COMPLETED")
    print("TOTAL API CALLS:",api_calls)
    print("====================================")